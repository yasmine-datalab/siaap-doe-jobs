import openpyxl.workbook
import xlrd
import openpyxl
import json
import os
import six
import io

rubrique_list = ["tr", "eq", "ou", "si", "vr"]
list_type_doc = ["bor", "cat", "cer", "cof", "com", "def", "doc", "enr", "fds", "fid",
              "fer", "fon", "lis", "man", "mop", "nor", "ndc", "pho", "paq", "pla",
              "pid", "pro", "prg", "sch", "syn", "ter"]

def extract_fulltext_from_sheet(sheet, nrows, ncols, engine="xlrd"):
    """
    Extrait le texte complet d'une feuille Excel sous forme de texte brut.

    Args:
        sheet: La feuille Excel.

    Returns:
        str: Texte brut représentant l'ensemble du contenu de la feuille.
    """
    fulltext = ""

    # Parcoure chaque ligne à partir de la deuxième
    for curr_row in range(1, nrows):
        print(curr_row)
        row_content = []
        for index_col in range(ncols):
            print(index_col)
            value = sheet.cell_value(curr_row, index_col) if engine=="xlrd" else sheet.cell(row=curr_row, column=index_col+1).value
            if value:
                if isinstance(value, (int, float)):
                    value = six.text_type(value)
                row_content.append(value)
        if row_content:
            fulltext += ' '.join(row_content) + '\n'

    return fulltext

def extract(filepath:str, filecontent:bytes):
    _, filename = os.path.split(filepath)
    name, extension = os.path.splitext(filename)

    rubrique = type_info = "Unknown"
    spec_doc = "Specifications" in filename

    # Séparation du nom du fichier en parties
    parts = name.split(' ')

    # Extraction de la rubrique et du type du fichier à partir du nom
    for part in parts:
        if part.lower() in rubrique_list:
            rubrique = part
        elif part.lower() in list_type_doc:
            type_info = part

    # Initialisation du dictionnaire des métadonnées
    metadata = {
            "name": name,
            "extension": [extension[1:]],
            "rubrique": rubrique,
            "type": type_info,
            "path": filepath,
            "spec_doc": spec_doc,
            "tags": [],
            "content": {
                "pages": []  
            }
        }
    
    try:
        # Essaie de charger avec xlrd
        if filepath.endswith((".xls")) and filecontent is not None:
            workbook = xlrd.open_workbook(file_contents=filecontent) 
            # Parcours chaque feuille du classeur
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                page_id = sheet_name.lower()  # Utilise le nom de la feuille comme ID de page
                fulltext = extract_fulltext_from_sheet(sheet, sheet.nrows, sheet.ncols)
                print(fulltext)

                # Ajoute le contenu à la clé "pages" du dictionnaire
                metadata["content"]["pages"].append( {
                    "id": page_id,
                    "fulltext": fulltext
                })
        if filepath.endswith((".xlsx")):
            workbook = openpyxl.load_workbook(io.BytesIO(filecontent))   # remplacer le filepath par un byteio
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                page_id = sheet_name.lower()  # Utilise le nom de la feuille comme ID de page
                fulltext = extract_fulltext_from_sheet(sheet, sheet.max_row, sheet.max_column, engine="openpyxl")
                print(fulltext)
                # Ajoute le contenu à la clé "pages" du dictionnaire
                metadata["content"]["pages"].append( {
                    "id": page_id,
                    "fulltext": fulltext
                })


        return metadata
    except Exception:
        pass

# Application:
# input_file_path = "DOE_SEM/SEM_Eng/Engineering - Technique/111540 bâtiment de production_EI/Specifications/1115 40 00x00-00 000000 Eq Fid 546-Conductivité rev.C.xls"

# def main(input_file_path):
#     metadata = extract(input_file_path)
#     json_text = json.dumps(metadata, ensure_ascii=False, indent=2)

#     # Sauvegarde le fichier json
#     with open('output.json', 'w', encoding='utf-8') as json_file:
#         json_file.write(json_text)
